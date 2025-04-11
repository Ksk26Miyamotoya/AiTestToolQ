"""
Excelレポート生成ユーティリティモジュール
"""
import os
import logging
from typing import Dict, Any
import gc
import tempfile
import shutil
import traceback
from datetime import datetime

try:
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    EXCEL_AVAILABLE = True
except ImportError:
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.utils import get_column_letter
    from PIL import Image as PILImage
    EXCEL_AVAILABLE = False


def generate_excel_report(results: Dict[str, Any], output_dir: str, config: Dict[str, Any]) -> str:
    """
    テスト結果をExcelレポートとして出力する
    
    Args:
        results: テスト結果
        output_dir: 出力ディレクトリ
        config: 設定情報
        
    Returns:
        Excelファイルのパス
    """
    # 独自のロガーを設定
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    if not logger.handlers:
        handler = logging.StreamHandler()
        handler.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    if not EXCEL_AVAILABLE:
        logger.error("openpyxlまたはPillowがインストールされていないため、Excelレポートを生成できません")
        return ""
    
    # 出力ディレクトリの作成
    result_dir = os.path.join(output_dir, "result")
    os.makedirs(result_dir, exist_ok=True)
    
    # スクリーンショットディレクトリの確認
    screenshot_dir = os.path.join(output_dir, "screenshot")
    
    # Excelファイルのパス
    excel_path = os.path.join(result_dir, "test_report.xlsx")
    
    # 一時ディレクトリの作成（画像処理用）
    temp_dir = tempfile.mkdtemp()
    temp_files = []
    
    # 日付フォーマットを統一（2025-04-04 03:05:04形式）
    date_format = "%Y-%m-%d %H:%M:%S"
    current_time = datetime.now().strftime(date_format)
    
    try:
        # ワークブックの作成
        wb = openpyxl.Workbook()
        
        # デフォルトシートの削除
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # サマリーシートの作成
        summary_sheet = wb.create_sheet("サマリー")
        
        # スタイルの設定
        header_bg_color = config.get("header_bg_color", "#4472C4")
        header_font_color = config.get("header_font_color", "#FFFFFF")
        alt_row_color = config.get("alt_row_color", "#E6F0FF")
        success_color = config.get("success_color", "#C6EFCE")
        failure_color = config.get("failure_color", "#FFC7CE")
        screenshot_title_cell_color = config.get("screenshot_title_cell_color", "#ffebcd")
        
        # レポート設定
        report_title = config.get("report_title", "テスト実行結果報告書")
        company_name = config.get("company_name", "")
        project_name = config.get("project_name", "")
        report_logo = config.get("report_logo", "")
        zoom_scale = config.get("zoom_scale", 50)
        include_timestamp = config.get("include_timestamp", True)
        timestamp_format = config.get("timestamp_format", "%Y-%m-%d %H:%M:%S")
        
        # 設定値のデバッグ出力
        logging.debug(f"Excelレポート設定:")
        logging.debug(f"  - レポートタイトル: {report_title}")
        logging.debug(f"  - 会社名: {company_name}")
        logging.debug(f"  - プロジェクト名: {project_name}")
        logging.debug(f"  - ロゴパス: {report_logo} (存在: {os.path.exists(report_logo) if report_logo else False})")
        logging.debug(f"  - ズーム倍率: {zoom_scale}%")
        logging.debug(f"  - タイムスタンプ表示: {include_timestamp}")
        logging.debug(f"  - タイムスタンプ形式: {timestamp_format}")
        
        # ヘッダースタイル
        header_fill = PatternFill(start_color=header_bg_color.replace("#", ""), end_color=header_bg_color.replace("#", ""), fill_type="solid")
        header_font = Font(bold=True, color=header_font_color.replace("#", ""))
        
        # 交互行スタイル
        alt_row_fill = PatternFill(start_color=alt_row_color.replace("#", ""), end_color=alt_row_color.replace("#", ""), fill_type="solid")
        
        # 成功/失敗スタイル
        success_fill = PatternFill(start_color=success_color.replace("#", ""), end_color=success_color.replace("#", ""), fill_type="solid")
        failure_fill = PatternFill(start_color=failure_color.replace("#", ""), end_color=failure_color.replace("#", ""), fill_type="solid")
        
        # スクリーンショットタイトルスタイル
        screenshot_title_fill = PatternFill(start_color=screenshot_title_cell_color.replace("#", ""), end_color=screenshot_title_cell_color.replace("#", ""), fill_type="solid")
        
        # 罫線スタイル
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # サマリーシートの作成（報告書形式）
        summary_sheet["A1"] = report_title
        summary_sheet["A1"].font = Font(size=16, bold=True)
        summary_sheet.merge_cells("A1:D1")
        summary_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        # 固定行で情報を表示する
        row = 2
        
        # デバッグ出力
        logging.debug(f"レポート設定: タイトル={report_title}, 会社名={company_name}, プロジェクト名={project_name}")
        
        # 会社名の設定（空でない場合のみ）
        if company_name and company_name.strip():
            summary_sheet[f"A{row}"] = "会社名:"
            summary_sheet[f"B{row}"] = company_name
            summary_sheet[f"A{row}"].font = Font(bold=True)
            logging.debug(f"会社名を設定しました: A{row}={company_name}")
            row += 1
            
        # プロジェクト名の設定（空でない場合のみ）
        if project_name and project_name.strip():
            summary_sheet[f"A{row}"] = "プロジェクト名:"
            summary_sheet[f"B{row}"] = project_name
            summary_sheet[f"A{row}"].font = Font(bold=True)
            logging.debug(f"プロジェクト名を設定しました: A{row}={project_name}")
            row += 1
            
        # タイムスタンプの追加（設定されている場合のみ）
        if include_timestamp:
            timestamp = datetime.now().strftime(timestamp_format if timestamp_format else "%Y-%m-%d %H:%M:%S")
            summary_sheet[f"A{row}"] = "実行日時:"
            summary_sheet[f"B{row}"] = timestamp
            summary_sheet[f"A{row}"].font = Font(bold=True)
            logging.debug(f"実行日時を設定しました: A{row}={timestamp}")
            row += 1
            
        # ロゴの追加（設定されている場合）
        if report_logo and os.path.exists(report_logo):
            try:
                # PILを使用して画像サイズを調整
                pil_img = PILImage.open(report_logo)
                # 適切なサイズに調整（例：幅200px）
                width, height = pil_img.size
                new_width = 200
                new_height = int(height * (new_width / width))
                
                # 一時ファイルに保存
                temp_logo = os.path.join(temp_dir, "temp_logo.png")
                pil_img.resize((new_width, new_height)).save(temp_logo)
                
                # Excelに挿入
                img = Image(temp_logo)
                summary_sheet.add_image(img, "D2")
                logging.debug(f"ロゴを追加しました: {report_logo} (サイズ: {new_width}x{new_height})")
            except Exception as e:
                logging.error(f"ロゴの追加に失敗しました: {str(e)}")
                logging.error(traceback.format_exc())
        
        # 空行を追加
        row += 1
        
        # 空行を追加
        row += 1
        
        # テスト結果サマリーのヘッダー
        summary_sheet[f"A{row}"] = "テスト結果サマリー"
        summary_sheet[f"A{row}"].font = Font(size=14, bold=True)
        summary_sheet.merge_cells(f"A{row}:D{row}")
        row += 1
        
        # 罫線の適用
        for r in range(row, row + 8):
            for col in range(1, 5):
                summary_sheet[f"{get_column_letter(col)}{r}"].border = thin_border
        
        # ヘッダー行のスタイル
        for col in range(1, 5):
            summary_sheet[f"{get_column_letter(col)}{row}"].fill = header_fill
            summary_sheet[f"{get_column_letter(col)}{row}"].font = header_font
        
        summary_sheet[f"A{row}"] = "項目"
        summary_sheet[f"B{row}"] = "内容"
        summary_sheet[f"C{row}"] = "項目"
        summary_sheet[f"D{row}"] = "内容"
        
        # 実行情報
        row += 1
        summary_sheet[f"A{row}"] = "実行日時"
        summary_sheet[f"B{row}"] = current_time
        
        row += 1
        summary_sheet[f"A{row}"] = "実行環境"
        summary_sheet[f"B{row}"] = config.get("browser", "chrome").capitalize()
        
        # アプリケーションユーザー情報
        app_users = results.get("app_users", [])
        if app_users:
            row += 1
            summary_sheet[f"A{row}"] = "実行ユーザー"
            summary_sheet[f"B{row}"] = ", ".join(app_users)
        
        # テスト結果
        total_sessions = len(results.get("sessions", []))
        success_sessions = sum(1 for session in results.get("sessions", []) if session.get("success", False))
        failure_sessions = total_sessions - success_sessions
        
        row_c4 = row - 2  # 実行日時の行に合わせる
        summary_sheet[f"C{row_c4}"] = "テスト総数"
        summary_sheet[f"D{row_c4}"] = f"{total_sessions} セッション"
        
        row_c5 = row - 1  # 実行環境の行に合わせる
        summary_sheet[f"C{row_c5}"] = "成功数"
        success_cell = summary_sheet[f"D{row_c5}"]
        success_cell.value = f"{success_sessions} セッション"
        success_cell.fill = success_fill
        
        row_c6 = row  # 実行ユーザーの行に合わせる
        summary_sheet[f"C{row_c6}"] = "失敗数"
        failure_cell = summary_sheet[f"D{row_c6}"]
        failure_cell.value = f"{failure_sessions} セッション"
        if failure_sessions > 0:
            failure_cell.fill = failure_fill
        
        # 成功率の行を追加（マージセルの問題を回避）
        row_c7 = row + 1
        summary_sheet[f"C{row_c7}"] = "成功率"
        success_rate_cell = summary_sheet[f"D{row_c7}"]
        success_rate = (success_sessions / total_sessions * 100) if total_sessions > 0 else 0
        success_rate_cell.value = f"{success_rate:.1f}%"
        if success_rate == 100:
            success_rate_cell.fill = success_fill
        elif success_rate < 80:
            success_rate_cell.fill = failure_fill
        
        # セッション一覧テーブル
        row_a9 = row_c7 + 2  # 成功率の2行下
        summary_sheet[f"A{row_a9}"] = "セッション一覧"
        summary_sheet[f"A{row_a9}"].font = Font(size=12, bold=True)
        summary_sheet.merge_cells(f"A{row_a9}:D{row_a9}")
        
        # セッション一覧ヘッダー
        headers = ["セッションID", "ユーザーID", "結果", "所要時間(秒)"]
        header_row = row_a9 + 1
        for i, header in enumerate(headers):
            cell = summary_sheet[f"{get_column_letter(i+1)}{header_row}"]
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # セッション一覧データ
        for i, session in enumerate(results.get("sessions", [])):
            data_row = header_row + i + 1
            
            # 交互行の背景色
            if i % 2 == 1:
                for col in range(1, len(headers) + 1):
                    summary_sheet[f"{get_column_letter(col)}{data_row}"].fill = alt_row_fill
            
            # セッション情報
            summary_sheet[f"A{data_row}"] = session.get("session_id", "unknown")
            summary_sheet[f"B{data_row}"] = session.get("user_id", "")
            
            # 結果
            result_cell = summary_sheet[f"C{data_row}"]
            result = session.get("success", False)
            result_cell.value = "成功" if result else "失敗"
            result_cell.fill = success_fill if result else failure_fill
            
            # 所要時間
            summary_sheet[f"D{data_row}"] = f"{session.get('duration', 0):.2f}"
            
            # 罫線
            for col in range(1, len(headers) + 1):
                summary_sheet[f"{get_column_letter(col)}{data_row}"].border = thin_border
                summary_sheet[f"{get_column_letter(col)}{data_row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        # 列幅の調整
        summary_sheet.column_dimensions["A"].width = 15
        summary_sheet.column_dimensions["B"].width = 25
        summary_sheet.column_dimensions["C"].width = 15
        summary_sheet.column_dimensions["D"].width = 15
        
        # セッション詳細シートとスクリーンショットシートを交互に作成
        for session in results.get("sessions", []):
            session_id = session.get("session_id", "unknown")
            
            # セッション詳細シートの作成
            # シート名の最大長は31文字
            sheet_name = f"セッション{session_id}"
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            session_sheet = wb.create_sheet(sheet_name)
            
            # ヘッダー
            session_sheet["A1"] = f"セッション{session_id} 詳細レポート"
            session_sheet["A1"].font = Font(size=16, bold=True)
            session_sheet.merge_cells("A1:F1")
            session_sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            
            # セッション情報テーブル
            # ヘッダー行のスタイル
            for col in range(1, 7):
                session_sheet[f"{get_column_letter(col)}3"].fill = header_fill
                session_sheet[f"{get_column_letter(col)}3"].font = header_font
                session_sheet[f"{get_column_letter(col)}3"].border = thin_border
                session_sheet[f"{get_column_letter(col)}3"].alignment = Alignment(horizontal="center", vertical="center")
            
            session_sheet["A3"] = "セッションID"
            session_sheet["B3"] = "ユーザーID"
            session_sheet["C3"] = "実行日時"
            session_sheet["D3"] = "結果"
            session_sheet["E3"] = "所要時間"
            session_sheet["F3"] = "テスト環境"
            
            # データ行
            session_sheet["A4"] = session_id
            session_sheet["B4"] = session.get("user_id", "")
            session_sheet["C4"] = current_time
            
            result_cell = session_sheet["D4"]
            result_cell.value = "成功" if session.get("success", False) else "失敗"
            result_cell.fill = success_fill if session.get("success", False) else failure_fill
            result_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            session_sheet["E4"] = f"{session.get('duration', 0):.2f} 秒"
            session_sheet["F4"] = config.get("browser", "chrome").capitalize()
            
            # 罫線とセルの配置
            for col in range(1, 7):
                session_sheet[f"{get_column_letter(col)}4"].border = thin_border
                session_sheet[f"{get_column_letter(col)}4"].alignment = Alignment(horizontal="center", vertical="center")
            
            # アクション一覧のヘッダー
            headers = ["操作ID", "操作タイプ", "対象要素", "入力値", "説明", "結果"]
            for i, header in enumerate(headers):
                cell = session_sheet[f"{get_column_letter(i+1)}5"]
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # アクション一覧
            actions = session.get("actions", [])
            if not actions:
                session_sheet["A6"] = "アクションデータがありません"
                session_sheet.merge_cells("A6:F6")
                session_sheet["A6"].alignment = Alignment(horizontal="center", vertical="center")
            else:
                for i, action in enumerate(actions):
                    row = i + 6
                    
                    # 交互行の背景色
                    if i % 2 == 1:
                        for col in range(1, len(headers) + 1):
                            session_sheet[f"{get_column_letter(col)}{row}"].fill = alt_row_fill
                    
                    # アクション情報
                    session_sheet[f"A{row}"] = action.get("操作ID", "")
                    session_sheet[f"B{row}"] = action.get("操作タイプ", "")
                    session_sheet[f"C{row}"] = action.get("対象要素", "")
                    session_sheet[f"D{row}"] = action.get("入力値", "")
                    session_sheet[f"E{row}"] = action.get("説明", "")
                    
                    # 結果
                    result_cell = session_sheet[f"F{row}"]
                    result = action.get("result", action.get("success", False))
                    result_cell.value = "成功" if result else "失敗"
                    result_cell.fill = success_fill if result else failure_fill
                    
                    # エラーメッセージがある場合は表示
                    if not result and action.get("error"):
                        error_row = row + 1
                        session_sheet[f"A{error_row}"] = "エラー:"
                        session_sheet[f"B{error_row}"] = action.get("error", "")
                        session_sheet.merge_cells(f"B{error_row}:F{error_row}")
                        session_sheet[f"B{error_row}"].fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
                        
                        # 罫線
                        for col in range(1, len(headers) + 1):
                            session_sheet[f"{get_column_letter(col)}{error_row}"].border = thin_border
                    
                    # 罫線
                    for col in range(1, len(headers) + 1):
                        session_sheet[f"{get_column_letter(col)}{row}"].border = thin_border
            
            # 列幅の調整
            session_sheet.column_dimensions["A"].width = 10
            session_sheet.column_dimensions["B"].width = 15
            session_sheet.column_dimensions["C"].width = 20
            session_sheet.column_dimensions["D"].width = 20
            session_sheet.column_dimensions["E"].width = 30
            session_sheet.column_dimensions["F"].width = 10
            
            # 対応するスクリーンショットシートを作成（セッションシートの直後に）
            # シート名の最大長は31文字
            screenshot_sheet_name = f"スクリーンショット{session_id}"
            if len(screenshot_sheet_name) > 31:
                screenshot_sheet_name = screenshot_sheet_name[:31]
            screenshots_sheet = wb.create_sheet(screenshot_sheet_name)
            
            # ヘッダー
            screenshots_sheet["A1"] = f"セッション{session_id} スクリーンショット"
            screenshots_sheet["A1"].font = Font(size=14, bold=True)
            screenshots_sheet.merge_cells("A1:C1")
            
            # スクリーンショットを収集
            screenshots = []
            
            # スクリーンショットディレクトリが存在するか確認
            if not os.path.exists(screenshot_dir):
                logging.warning(f"スクリーンショットディレクトリが存在しません: {screenshot_dir}")
                screenshots_sheet["A2"] = "スクリーンショットディレクトリが存在しません"
                continue
                
            # セッションIDに関連するスクリーンショットを検索
            session_dir = os.path.join(screenshot_dir, f"session_{session_id}")
            
            if not os.path.exists(session_dir):
                logging.warning(f"セッションディレクトリが存在しません: {session_dir}")
                screenshots_sheet["A2"] = "セッションディレクトリが存在しません"
                continue
            
            # セッションディレクトリ内のすべてのスクリーンショットを再帰的に検索
            for root, _, files in os.walk(session_dir):
                for file in files:
                    if file.endswith('.png'):
                        # Excel出力フラグをチェック
                        marker_file = os.path.join(root, file + ".excel")
                        if os.path.exists(marker_file):  # Excel出力フラグがある場合のみ追加
                            screenshots.append(os.path.join(root, file))
                            logging.debug(f"Excel出力用スクリーンショット追加: {os.path.join(root, file)}")
                        else:
                            logging.debug(f"Excel出力用マーカーファイルがないためスキップ: {marker_file}")
            
            if not screenshots:
                logging.warning(f"セッション {session_id} のスクリーンショットが見つかりませんでした")
                screenshots_sheet["A2"] = "スクリーンショットが見つかりませんでした"
                continue
            
            # スクリーンショットをアクションIDでグループ化
            screenshot_pairs = []
            
            # スクリーンショットのパスからアクションIDを抽出する関数
            def extract_action_id(path):
                # パスからファイル名を取得
                filename = os.path.basename(path)
                # ファイル名からアクションIDを抽出
                parts = filename.split('_')
                if len(parts) >= 3:
                    if parts[0] in ['before', 'after', 'error']:
                        return parts[1]
                return None
            
            # スクリーンショットをアクションIDでグループ化
            action_screenshots = {}
            for screenshot in screenshots:
                action_id = extract_action_id(screenshot)
                if action_id:
                    # ファイル名からセッションIDを抽出
                    filename = os.path.basename(screenshot)
                    parts = filename.split('_')
                    screenshot_session_id = None
                    
                    # ファイル名からセッションIDを抽出
                    for i, part in enumerate(parts):
                        if part == "session" and i + 1 < len(parts):
                            try:
                                screenshot_session_id = int(parts[i + 1])
                                break
                            except ValueError:
                                pass
                    
                    # セッションIDが一致する場合のみ処理
                    if screenshot_session_id == session_id:
                        # 各アクションIDに対して複数のスクリーンショットを保持
                        if action_id not in action_screenshots:
                            action_screenshots[action_id] = {'before': None, 'after': None, 'error': None}
                        
                        if filename.startswith('before'):
                            action_screenshots[action_id]['before'] = screenshot
                        elif filename.startswith('after'):
                            action_screenshots[action_id]['after'] = screenshot
                        elif filename.startswith('error'):
                            action_screenshots[action_id]['error'] = screenshot
                    else:
                        logging.debug(f"セッションID不一致のためスキップ: {filename}, 期待: {session_id}, 実際: {screenshot_session_id}")
            
            
            # グループ化したスクリーンショットをリストに変換
            for action_id, pair in action_screenshots.items():
                screenshot_pairs.append({
                    'action_id': action_id,
                    'before': pair['before'],
                    'after': pair['after'],
                    'error': pair['error']
                })
            
            # アクションIDでソート
            screenshot_pairs.sort(key=lambda x: int(x['action_id']) if x['action_id'].isdigit() else float('inf'))
            
            # スクリーンショットの表示
            row = 3
            for idx, pair in enumerate(screenshot_pairs):
                # アクションIDの表示
                action_id = pair["action_id"]
                action_info = next((a for a in actions if str(a.get("操作ID", "")) == action_id), None)
                
                action_description = ""
                if action_info:
                    action_description = action_info.get('説明', '')
                
                screenshots_sheet[f"A{row}"] = f"アクション {action_id}: {action_description}"
                screenshots_sheet[f"A{row}"].fill = screenshot_title_fill
                screenshots_sheet[f"A{row}"].font = Font(bold=True)
                screenshots_sheet.merge_cells(f"A{row}:C{row}")

                # 列幅の調整
                screenshots_sheet.column_dimensions["A"].width = 20
                screenshots_sheet.column_dimensions["B"].width = 143
                screenshots_sheet.column_dimensions["C"].width = 143
                
                # ラベル行
                row += 1
                
                # 画像の有無に応じてラベルを設定
                has_before = pair["before"] is not None
                has_after = pair["after"] is not None or pair["error"] is not None
                
                # Beforeラベルの設定（画像がある場合のみ）
                if has_before:
                    screenshots_sheet[f"B{row}"] = "Before:"
                
                # After/Errorラベルの設定（画像がある場合のみ）
                if has_after:
                    # 表示する列を決定（Beforeがない場合はB列、ある場合はC列）
                    after_col = "B" if not has_before else "C"
                    
                    # エラー画像がある場合は「Error:」、そうでなければ「After:」を表示
                    if pair["error"]:
                        error_cell = screenshots_sheet[f"{after_col}{row}"]
                        error_cell.value = "Error:"
                        # エラーラベルのセルを赤色に設定
                        error_cell.fill = PatternFill(start_color=failure_color.replace("#", ""), 
                                                     end_color=failure_color.replace("#", ""), 
                                                     fill_type="solid")
                        
                        # 背景色の明るさを判断して文字色を設定
                        # 16進数の色コードをRGB値に変換
                        r = int(failure_color.replace("#", "")[0:2], 16)
                        g = int(failure_color.replace("#", "")[2:4], 16)
                        b = int(failure_color.replace("#", "")[4:6], 16)
                        
                        # 明るさの計算（一般的な輝度計算式）
                        brightness = (r * 299 + g * 587 + b * 114) / 1000
                        
                        # 明るさが128以上なら黒、それ以下なら白
                        font_color = "000000" if brightness >= 128 else "FFFFFF"
                        error_cell.font = Font(bold=True, color=font_color)
                    else:
                        screenshots_sheet[f"{after_col}{row}"] = "After:"
                
                # 画像行
                row += 1
                
                # Before画像の処理
                if pair["before"]:
                    try:
                        # 一時ファイルのパスを作成
                        temp_file_before = os.path.join(temp_dir, f"temp_before_{session_id}_{action_id}_{idx}.png")
                        temp_files.append(temp_file_before)
                        
                        # 画像をそのままコピー（リサイズなし）
                        shutil.copy(pair["before"], temp_file_before)
                        
                        # Excelに挿入
                        excel_img = Image(temp_file_before)
                        
                        # 画像の位置を調整
                        screenshots_sheet.add_image(excel_img, f"B{row}")
                        
                        # デバッグログ
                        logging.debug(f"Before画像を挿入しました: {pair['before']} -> {temp_file_before}")
                            
                    except Exception as e:
                        logging.error(f"Before画像の挿入に失敗しました: {str(e)}, パス: {pair['before']}")
                        screenshots_sheet[f"B{row}"] = f"画像の挿入に失敗: {str(e)}"
                
                # After/Error画像の処理
                after_image = pair["error"] if pair["error"] else pair["after"]
                if after_image:
                    try:
                        # 一時ファイルのパスを作成
                        temp_file_after = os.path.join(temp_dir, f"temp_after_{session_id}_{action_id}_{idx}.png")
                        temp_files.append(temp_file_after)
                        
                        # 画像をそのままコピー（リサイズなし）
                        shutil.copy(after_image, temp_file_after)
                        
                        # Excelに挿入
                        excel_img = Image(temp_file_after)
                        
                        # 画像の位置を調整（Before画像がない場合はB列に表示）
                        after_col = "B" if not pair["before"] else "C"
                        screenshots_sheet.add_image(excel_img, f"{after_col}{row}")
                        
                        # デバッグログ
                        logging.debug(f"After/Error画像を挿入しました: {after_image} -> {temp_file_after} (列: {after_col})")
                            
                    except Exception as e:
                        logging.error(f"After/Error画像の挿入に失敗しました: {str(e)}, パス: {after_image}")
                        after_col = "B" if not pair["before"] else "C"
                        screenshots_sheet[f"{after_col}{row}"] = f"画像の挿入に失敗: {str(e)}"
                
                # 次のアクションのための行を確保（50行の間隔を開ける）
                row += 50
        
        # ワークブックの保存
        # スクリーンショットシートの表示倍率を設定
        for sheet_name in wb.sheetnames:
            if "スクリーンショット" in sheet_name:
                sheet = wb[sheet_name]
                sheet.sheet_view.zoomScale = zoom_scale  # 設定から表示倍率を取得
                logging.debug(f"シート '{sheet_name}' の表示倍率を{zoom_scale}%に設定しました")
        
        wb.save(excel_path)
        logging.info(f"Excelレポートを保存しました: {excel_path}")
        
        # 明示的にメモリを解放
        del wb
        gc.collect()
        
        return excel_path
    
    except Exception as e:
        logging.error(f"Excelレポートの生成に失敗しました: {str(e)}")
        logging.error(traceback.format_exc())
        return ""
    
    finally:
        # 一時ファイルの削除
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        # 一時ディレクトリの削除
        if os.path.exists(temp_dir):
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
